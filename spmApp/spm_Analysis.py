import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QComboBox, QTextEdit, QFileDialog, QLabel, QLineEdit, QGroupBox
from PyQt5.QtGui import QPixmap, QIcon  # Import QIcon
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QIcon 
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QDesktopServices
import pandas as pd
import numpy as np
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import pdfplumber
import os
from functools import partial
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import time  # Simulating delay
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QFileDialog, QTextEdit, QLabel, QLineEdit, QGroupBox)
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt
import pandas as pd
import sys
from functools import partial
from fpdf import FPDF
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.dates as mdates
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go
import warnings


def cms_data_sheet(self):
    window.text_box.clear()
    window.text_box.append("Please Wait !!!!, Updating Crew & CLI Details........")
    QApplication.processEvents()  # Force UI update
   # Connect to Google
    scope = ["https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.file",
            "https://www.googleapis.com/auth/drive"]

    credentials = ServiceAccountCredentials.from_json_keyfile_name("gs_credentials.json", scope)
    client = gspread.authorize(credentials)

    try:
        # Open the Google Sheet by ID
        sheet = client.open_by_key('1f41XlaUwmldxK_3gx2Wk55XSgV1mAW2T-QFcrqSG38M').sheet1  # Replace 'your_sheet_id_here' with the actual sheet ID

        # Get all records from the sheet
        records = sheet.get_all_records()

        # Convert to a DataFrame
        df = pd.DataFrame(records)

        if os.path.exists('CMS_Data.xlsx'):
            os.remove('CMS_Data.xlsx')

        # Save the filtered data to an Excel file named "Filtered_CMS_Data.xlsx" in the root folder
        df.to_excel('CMS_Data.xlsx', index=False)
        window.text_box.clear()
        window.text_box.append("All Set !!!!, Please Proceed........")
        QApplication.processEvents()  # Force UI update
        
        print("Data successfully saved to 'CMS_Data.xlsx'")
        
    except gspread.exceptions.SpreadsheetNotFound as e:
        print("Data not found:", e)
    except Exception as ex:
        print("An error occurred:", ex)
   

def FAQ(self):
    url = QUrl("https://docs.google.com/presentation/d/1F-lzQE7awUBnW_RNFO0was9-UGXmBzjyUFCq6-mg6lA/present?slide=id.p")
    QDesktopServices.openUrl(url)
    window.text_box.clear()
    window.text_box.append("Share Screenshot to 'binaykumarlucky@gmail.com' in case of error found")
    QApplication.processEvents()  # Force UI update

def launch_streamlit_app():
    # Popen(["streamlit", "run", "./graph.py"])
    url = QUrl("https://loco-pilot-driving-skill-analysis.streamlit.app/?embed_options=light_theme,show_colored_line")
    QDesktopServices.openUrl(url)


# ===========MEDHA=========================================================================================================
def medha(self):
    window.text_box.clear()
    window.text_box.append("File Processing Please Wait........")
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getOpenFileName(None, "Upload Medha File", "", "Text Files (*.txt);;Excel Files (*.xlsx *.xls)", options=options)
    if file_path:
        encodings_to_try = ['utf-8', 'latin-1', 'utf-16', 'windows-1252']  # Add more encodings if needed
        print("Selected file:", file_path)
        try:
            # Read the file into a pandas DataFrame
            if file_path.endswith('.txt'):
                # Try different encodings to read text file
                for encoding in encodings_to_try:
                    try:
                        df = pd.read_csv(file_path, delimiter='\t', encoding=encoding)
                        print("File successfully read with encoding:", encoding)
                        # Process the DataFrame or do further operations
                        break  # Exit the loop if file is successfully read
                    except Exception as e:
                        print(f"Failed to read with encoding {encoding}: {str(e)}")
                else:
                    print("Unable to read the file with any encoding")

            else:
                # Read Excel file
                df = pd.read_excel(file_path)
                print("File is an Excel file. Reading as Excel.")

            # Process the DataFrame or do further operations
        except Exception as e:
            window.text_box.clear()
            window.text_box.append(f"An error occurred while processing the file: {str(e)}")
            QApplication.processEvents()  # Force UI update
            print(f"An error occurred while processing the file: {str(e)}")
    # Define the new header names  .....................................................................................
        new_header_names = ['New Header 1']
        df.columns = new_header_names

        # Split the text in Column 1 using "|" delimiter
        split_values = df['New Header 1'].str.split("|", expand=True)
        df['Date'] = split_values[0]
        df['Time'] = split_values[1]
        df['Speed'] = split_values[2]
        df['Distance'] = split_values[3]
        df['Column 5'] = ''
        df['Column 6'] = ''
        df['Column 7'] = ''
        df['Column 8'] = ''
        df['Column 9'] = ''
        df['Column 10'] = ''
        df['Column 11'] = ''
        
        # Delete unnecessary columns
        columns_to_delete = ['Column 5', 'Column 6', 'Column 7', 'Column 8', 'Column 9', 'Column 10', 'Column 11']
        df.drop(columns=columns_to_delete, inplace=True)

        window.text_box.clear()
        window.text_box.append("25% Processing Done, Plase Wait..........")
        QApplication.processEvents()  # Force UI update

        # Cut and paste Column 1 cell value to Column 5 if it contains "Driver"
        df.loc[df['Date'].str.contains('Driver', na=False), 'Column 5'] = df['Date']

        # Copy down cell value of Column 5 if Column 4 data is not null
        df['Column 5'] = df['Column 5'].ffill()

        # Delete the Entire row with garbage value (Condition applied "Where Col 3 is null")
        df = df.dropna(subset=['Speed'])
        df.reset_index(drop=True, inplace=True)

        # Delete Column 1 which is useless now
        df = df.drop("New Header 1", axis=1)
        
        # Split values in Column 5 using ":"
        split_values = df['Column 5'].str.split(":", expand=True)
        df['Column 6'] = split_values[0]
        df['Column 7'] = split_values[1]
        df['Column 8'] = split_values[2]
        df['Column 9'] = split_values[3]
        # Delete extra text from each row in Column 7, 8, 9 and Remove White Space Using ".str.strip()"
        df['CMS_ID'] = df['Column 7'].str.replace('Train No', '').str.strip()
        df['Train_No'] = df['Column 8'].str.replace('Locono', '').str.strip()
        df['Loco_No'] = df['Column 9'].str.replace('Spd Limit', '').str.strip()
# Function to clean and convert the text to time format ..........................................................................................
        def clean_and_convert_time(time_str):
            # Remove leading and trailing white spaces
            time_str = time_str.strip()
            # Use regular expression to find and extract time components
            match = re.match(r'^(\d+):(\d+):(\d+)$', time_str)
            if match:
                # If the regex pattern matches, extract components and format the time string
                hour, minute, second = match.groups()
                return f"{hour.zfill(2)}:{minute.zfill(2)}:{second.zfill(2)}"
            else:
                # If the regex pattern doesn't match, return None (or handle as desired)
                return None
# Apply the function to the 'Time' column and create a new column 'Formatted_Time' ..................................................................
        df['Time'] = df['Time'].apply(clean_and_convert_time)
        
        # Delete unnecessary columns
        columns_to_delete = ['Column 5', 'Column 6', 'Column 7', 'Column 8', 'Column 9']
        df.drop(columns=columns_to_delete, inplace=True)    
# Convert certain columns to numeric format .................................................................................................
        numeric_columns = ['Speed', 'Distance','Loco_No']
        df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')

        print("Basic Columns Done")
        window.text_box.clear()
        window.text_box.append("50% Processing Done, Plase Wait..........")
        QApplication.processEvents()  # Force UI update
# Cumulate the value of 'Distance' based on the Start & Stop ...................................................................................
        cum_distance = 0
        for index, row in df.iterrows():
            if row['Speed'] == 0:
                cum_distance = 0
            else:
                cum_distance += row['Distance']
            df.loc[index, 'Cum_Dist_Run'] = cum_distance
# Cumulate the value of 'Distance' based on the CMS ID ...................................................................................
        cum_distance = 0
        prev_cms_id = None
        for i in range(len(df)):
            if df.loc[i, 'CMS_ID'] != prev_cms_id:
                cum_distance = 0
                prev_cms_id = df.loc[i, 'CMS_ID']
            else:
                cum_distance += df.loc[i, 'Distance']
            df.loc[i, 'Cum_Dist_LP'] = cum_distance
        print("Cum_Dist_LP Done")

# Create a new column 'Run_No' based on the conditions ...................................................................................
        df['Run_No'] = 1  
        cms_change = df['CMS_ID'] != df['CMS_ID'].shift()
        distance_reset = df['Cum_Dist_Run'] < df['Cum_Dist_Run'].shift()
        group = cms_change.cumsum()
        df['Run_No'] = distance_reset.groupby(group).cumsum() + 1
        print("Run_No Done")

# Remove entire rows where 'CMS ID' contains 0........................................................................
        # df = df[df['CMS_ID'] != '0']

# Insert a new column 'Run_Sum' with the sum of 'Distance' per 'Run_No' ....................................................................
        df['Run_Sum'] = df.groupby(['Run_No','CMS_ID'])['Distance'].transform('sum')

# Deduct values from 'Run_Sum' to 'Distance' column ........................................................................................
        df['Rev_Dist'] = df['Run_Sum'] - df['Cum_Dist_Run']
# Create a new column 'Pin_Point' with value "10 Meters" for the rows closest to 10 in each 'Run_No' group .........................................................
        df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 10).idxmin()) == df.index, '10 Meters', '')
        # Update 'Pin_Point' column with other values
        df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 250).idxmin()) == df.index, '250 Meters', df['Pin_Point'])
        df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 500).idxmin()) == df.index, '500 Meters', df['Pin_Point'])
        df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 1000).idxmin()) == df.index, '1000 Meters', df['Pin_Point'])
        print("Pin Point Done")
# Add BFT Column ......................................................................................................................
        df['Speed_shift'] = df['Speed'].shift(-1)
        unique_cms_ids = set()
        def add_bft(row):
            if row['Cum_Dist_LP'] < 10000 and 15 <= row['Speed'] <= 18 and row['Speed'] > row['Speed_shift']:
                if row['CMS_ID'] not in unique_cms_ids:
                    unique_cms_ids.add(row['CMS_ID'])
                    return 'BFT'
            return ''
        df['BFT'] = df.apply(add_bft, axis=1)
        print("BFT Column Done")
# Add BPT Column...............................................................................................................................
        unique_cms_ids = set()
        def add_bpt(row):
            if row['Cum_Dist_LP'] < 10000 and 40 <= row['Speed'] <= 90 and row['Speed'] > row['Speed_shift']:
                if row['CMS_ID'] not in unique_cms_ids:
                    unique_cms_ids.add(row['CMS_ID'])
                    return 'BPT'
            return ''
        df['BPT'] = df.apply(add_bpt, axis=1)

        # Reset shift column for comparison
        df['Speed_shift'] = df['Speed'].shift(1)
        print("BPT Column Done")
# BFT_END .............................................................................................................................
        def get_bft_end(df):
            df['BFT_END'] = ''
            for cms_id, group in df.groupby('CMS_ID'):
                group = group.reset_index()
                bft_idx = group[group['BFT'] == 'BFT'].index
                if not bft_idx.empty:
                    bft_idx = bft_idx[0]
                    for i in range(bft_idx + 1, len(group)):
                        if group.loc[i, 'Speed'] > group.loc[i, 'Speed_shift'] and 0 <= group.loc[i, 'Speed'] <= 9 and group.loc[i, 'Cum_Dist_LP'] < 10000:
                            end_idx = i - 1
                            if end_idx >= 0:
                                df.loc[group.loc[end_idx, 'index'], 'BFT_END'] = 'BFT_END'
                            break
            return df
        print("BFT_END Column Done")
# BPT_END ..............................................................................................................................
        def get_bpt_end(df):
            df['BPT_END'] = ''
            for cms_id, group in df.groupby('CMS_ID'):
                group = group.reset_index()
                bpt_idx = group[group['BPT'] == 'BPT'].index
                if not bpt_idx.empty:
                    bpt_idx = bpt_idx[0]
                    for i in range(bpt_idx + 1, len(group)):
                        if group.loc[i, 'Speed'] > group.loc[i, 'Speed_shift'] and 0 <= group.loc[i, 'Speed'] <= 45 and group.loc[i, 'Cum_Dist_LP'] < 10000:
                            end_idx = i - 1
                            if end_idx >= 0:
                                df.loc[group.loc[end_idx, 'index'], 'BPT_END'] = 'BPT_END'
                            break
            return df
        print("BPT_END Column Done")
        
        # Apply end marking functions
        df = get_bft_end(df)
        df = get_bpt_end(df)

        # Drop Speed_shift
        df.drop(columns=['Speed_shift'], inplace=True)
        window.text_box.clear()
        window.text_box.append("75% Processing Done, Plase Wait..........")
        QApplication.processEvents()  # Force UI update

# # Adding a new column 'BFT_BPT' ............................................................................................................................
        df['BFT_BPT'] = df.apply(lambda row: 
            (str(row['BFT']) if pd.notna(row['BFT']) and row['BFT'] != '' else '') +
            (' ' + str(row['BPT']) if pd.notna(row['BPT']) and row['BPT'] != '' else '') +
            (' ' + str(row['BFT_END']) if pd.notna(row['BFT_END']) and row['BFT_END'] != '' else '') +
            (' ' + str(row['BPT_END']) if pd.notna(row['BPT_END']) and row['BPT_END'] != '' else ''),
            axis=1
            ).str.strip()

        print(df.head())
# Adding a new column 'Crew Name' , 'CLI Name', 'Desig'............................................................................................................................
        try:
            cms_file_path = 'CMS_Data.xlsx'
            cms_df = pd.read_excel(cms_file_path)

            # Your existing mapping logic using the CMS data
            df['Crew_Name'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[1]])
            df['Nom_CLI'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[4]])
            df['Desig'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[2]])

        except FileNotFoundError:
            # File not found handling: Create columns with null values
            df['Crew_Name'] = None
            df['Nom_CLI'] = None
            df['Desig'] = None


    # Create a New Sheet "Report" and Push Data to New Sheet .................................................................................................
        df_results = df[df['Pin_Point'] !=''][['Date', 'Time', 'CMS_ID', 'Train_No', 'Loco_No','Run_No', 'Speed', 'Rev_Dist', 'Pin_Point']]
        df_bpt = df[df['BFT_BPT'] !=''][['Date', 'Time', 'CMS_ID', 'Train_No', 'Loco_No','Run_No', 'Speed', 'BFT_BPT']]
        print("Almost Done")
        window.text_box.clear()
        window.text_box.append("100% Processing Done, Plase Wait..........")
        QApplication.processEvents()  # Force UI update


        # Rearrange the Columns
        df = df[["Date", "Time", "Speed", "Distance", "CMS_ID", "Train_No", "Loco_No", "Crew_Name","Nom_CLI","Desig","BFT","BFT_END","BPT","BPT_END","BFT_BPT",
                 "Cum_Dist_Run","Cum_Dist_LP","Run_No","Run_Sum","Rev_Dist","Pin_Point"]]

 # Save Data to Excel
        options = QFileDialog.Options() 
        save_path, _ = QFileDialog.getSaveFileName(None, "Save Processed Data", "", "Excel Files (*.xlsx)", options=options)

        # Check if a save location was selected
        if save_path:
            # Save the processed data to a single sheet "DataBase"
            df.to_excel(save_path, sheet_name="DataBase", index=False, engine="openpyxl")
            print("File saved successfully.")
            window.text_box.clear()
            window.text_box.append("Done Successfully!!!, Ready For Visual Analysis or Quick Report...")
            QApplication.processEvents()  # Force UI update
        else:
            print("Save canceled.")
            window.text_box.clear()
            window.text_box.append("Process Aborted...!!!!")
            QApplication.processEvents()  # Force UI update
        
        # print(df)
        return df


# ========TELPRO PDF=======================================================================================================================================================
def telpro():
    window.text_box.clear()
    window.text_box.append("Telpro File Processing Please Wait........")
    QApplication.processEvents()  # Force UI update
    options = QFileDialog.Options()
    fileName, _ = QFileDialog.getOpenFileName(None, "Open PDF File", "", "PDF Files (*.pdf);;All Files (*)", options=options)
    
    if fileName:
        process_pdf(fileName)  # Pass fileName as an argument to extractDataFromPDF
    else:
        window.text_box.clear()
        window.text_box.append("No file selected.")
        QApplication.processEvents()  # Force UI update


def process_pdf(pdf_path):
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    all_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
                window.text_box.clear()
                window.text_box.append(f"Reading Page {page_num} of {len(pdf.pages)}...")
                QApplication.processEvents()  # Force UI update
                time.sleep(0.1)  # Simulating processing delay
                # print(f"Processing Page {page_num}...")
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                all_data.extend(table)  # Add all rows without assuming a fixed structure
                window.text_box.clear()
                window.text_box.append("PDF Under Process, Plase Wait..........")
                QApplication.processEvents()  # Force UI update

            text = page.extract_text()
            if text:
                lines = text.split('\n')
                for line in lines:
                    data = line.split()  # Example split, adjust as needed
                    all_data.append(data)

    # Determine the maximum number of columns in the dataset
    max_cols = max(len(row) for row in all_data)

    # Create a DataFrame without fixing the column names
    df = pd.DataFrame(all_data, columns=[f"Column_{i+1}" for i in range(max_cols)])
    print("Data Read by System")
    window.text_box.clear()
    window.text_box.append("PDF Under Process, Plase Wait..........")
    QApplication.processEvents()  # Force UI update
# -------------------------------------------------------------------------------------------------------

    # Create a Column "Status" based on data occurrences of "RUN" in rows
    for col in df.columns:
        if (df[col] == "RUN").sum() >= 10:
            df.rename(columns={col: "Status"}, inplace=True)

    # Insert new column "data" after "Column_4"
    df.insert(df.columns.get_loc("Column_4") + 1, "data", "")

    # Set the required column sequence
    column_order = [
        "Column_1", "Column_2", "Column_3", "Column_4", "data","Column_6", "Column_7", 
        "Column_8", "Column_9","Column_10", "Column_11", "Status"]
    df = df[column_order]    # Reorder DataFrame

    print("Status & Data Column Created and Reordered")

# -------------------------------------------------------------------------------------------------------
    # Identify columns to clear (all except the required ones)
    columns_to_clear = [col for col in df.columns if col not in ["Column_1", "Column_2", "Column_3", "Column_4", "Status"]]
    # Clear data from unwanted columns by setting values to empty strings or NaN
    df[columns_to_clear] = ""
# -------------------------------------------------------------------------------------------------------
    # Copy Column_1 data to Column_5 for those rows who has Concatenated Value with Driver ID:-
    if not df.empty and "Column_1" in df and "data" in df:  # Identify rows where Column_1 contains "Driver"
        driver_rows = df["Column_1"].astype(str).str.contains("Driver ID:-", na=False, regex=True)
        df.loc[driver_rows, "data"] = df.loc[driver_rows, "Column_1"]
    if "data" not in df.columns:  # Check if "Column_1" exists before filtering
        print("Error: data not found in extracted data.")
        return
# -------------------------------------------------------------------------------------------------------    
    def extract_values(row):
        match = re.search(
            r"Driver ID:-\s*([\S]+)\s*Train No:-\s*([\S ]+?)\s*Tr Load\(Ton\):-\s*([\S]+)\s*Loco No:-\s*([\S]+)\s*Wheel Dia\(mm\):-\s*([\S]+)\s*SpeedLimit\(kmph\):-\s*([\S]+)",
            str(row),
            re.DOTALL  # Enables multi-line matching
        )
        if match:
            return list(match.groups())  # Convert tuple to list
        return [None] * 6  # Ensure every row has exactly 6 values

    # Apply function to DataFrame
    df[["Column_6", "Column_7", "Column_8", "Column_9", "Column_10", "Column_11"]] = df["data"].apply(lambda x: pd.Series(extract_values(x)))
    print("Splittings Done")
# -------------------------------------------------------------------------------------------------------
     # Reaaarnge Columns
    desired_order = [
        "Column_1", "Column_2", "Column_3", "Column_4", "Status", "data", 
        "Column_6", "Column_7", "Column_8", "Column_9", "Column_10", "Column_11"
    ]
    df = df[desired_order]
# ----------------------------------------------------------------------------------------------
    # Columns to apply the fill operation
    columns_to_fill = ["Column_6", "Column_7", "Column_8", "Column_9", "Column_10", "Column_11"]

    # Find indices where 'Status' is "CHANGED"
    changed_idx = df.index[df["Status"] == "CHANGED"].tolist()
    start = 0
    for i in range(len(changed_idx)):
        # Fill down from the last valid data point until "CHANGED"
        df.loc[start:changed_idx[i]-1, columns_to_fill] = df.loc[start:changed_idx[i]-1, columns_to_fill].ffill()

        # After "CHANGED", find the next valid data and fill backwards
        next_valid = df[columns_to_fill].iloc[changed_idx[i]+1:].first_valid_index()
        if next_valid is not None:
            df.loc[changed_idx[i]+1:next_valid, columns_to_fill] = df.loc[changed_idx[i]+1:next_valid, columns_to_fill].bfill()

        # Move to next section
        start = changed_idx[i] + 1

    # Fill down the last segment after the last "CHANGED"
    df.loc[start:, columns_to_fill] = df.loc[start:, columns_to_fill].ffill()

# -------------------------------------------------------------------------------------------------------
    # Filter: Keep rows where Column_1 contains "Driver ID:-" or "/2"
    df = df[df["Status"].astype(str).str.contains("RUN|STOP", na=False, regex=True)]

     # Reaaarnge Columns
    desired_order = [
        "Column_1", "Column_2", "Column_3", "Column_4", 
        "Column_6", "Column_7", "Column_8", "Column_9", "Column_10", "Column_11", "Status", "data"
    ]
    df = df[desired_order]

    # Rename first four columns
    df.columns = ["Date", "Time", "Distance", "Speed", "CMS_ID", "Train_No","Column_8","Loco_No"] + list(df.columns[8:])
    # Rearrange the Columns
    df = df[["Date", "Time", "Speed", "Distance", "CMS_ID", "Train_No", "Loco_No", "Column_8","Status", "data"]]

    print("Analyzing PDF File, Please Wait.........")
    window.text_box.clear()
    window.text_box.append("Saving File to Excel!!!")

 # # Columns to be deleted
    columns_to_delete = ["Column_8","data", "Status"]
     # Ensure the columns exist before dropping
    df = df.drop(columns=[col for col in columns_to_delete if col in df.columns], errors="ignore")
   
    print("Basic Columns Done")
    window.text_box.clear()
    window.text_box.append("Generating Excel- 25 %, Plase Wait..........")
    QApplication.processEvents()  # Force UI update

    # Trim white spaces from all string columns
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    # Convert certain columns to numeric format .................................................................................................
    numeric_columns = ['Speed', 'Distance', 'Loco_No']
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
    # Drop rows where "Speed" or "Distance" is NaN (non-numeric values)
    df = df.dropna(subset=["Speed", "Distance"])

# Function to clean and convert the text to time format ..........................................................................................
    def clean_and_convert_time(time_str):
        # Remove leading and trailing white spaces
        time_str = time_str.strip()
        # Use regular expression to find and extract time components
        match = re.match(r'^(\d+):(\d+):(\d+)$', time_str)
        if match:
            # If the regex pattern matches, extract components and format the time string
            hour, minute, second = match.groups()
            return f"{hour.zfill(2)}:{minute.zfill(2)}:{second.zfill(2)}"
        else:
            # If the regex pattern doesn't match, return None (or handle as desired)
            return None
    df['Time'] = df['Time'].apply(clean_and_convert_time)

# Cumulate the value of 'Distance' based on the Start & Stop .................................................................................
    # Convert 'Distance' column to numeric, coercing errors to NaN
    df['Distance'] = pd.to_numeric(df['Distance'], errors='coerce')

    # Drop rows where 'Distance' is NaN
    df = df.dropna(subset=['Distance'])
    df['Cum_Dist_Run'] = df.groupby(df['Speed'].eq(0).cumsum())['Distance'].cumsum()

# Create a new column 'Run_No' based on the conditions ...................................................................................
    # Initialize 'Run_No' as 1
    df['Run_No'] = 1  
    # Identify when 'CMS_ID' changes
    cms_change = df['CMS_ID'] != df['CMS_ID'].shift()
    # Identify when 'Cum_Dist_Run' decreases within the same CMS_ID
    distance_reset = df['Cum_Dist_Run'] < df['Cum_Dist_Run'].shift()
    # Create a new group every time CMS_ID changes
    group = cms_change.cumsum()
    # Compute Run_No within each CMS_ID group, starting from 1
    df['Run_No'] = distance_reset.groupby(group).cumsum() + 1

# Cumulate the value of 'Distance' based on the Start & Stop .................................................................................
    df['Cum_Dist_Run'] = df.groupby(df['Speed'].eq(0).cumsum())['Distance'].cumsum()

# Cumulate the value of 'Distance' based on the LP .................................................................................
    if {'Speed', 'CMS_ID', 'Distance'}.issubset(df.columns):
        df['Cum_Dist_LP'] = df.groupby('CMS_ID')['Distance'].cumsum()

# Insert a new column 'Run_Sum' with the sum of 'Distance' per 'Run_No' ....................................................................
    df['Run_Sum'] = df.groupby(['Run_No','CMS_ID'])['Distance'].transform('sum')

# Deduct values from 'Run_Sum' to 'Distance' column ........................................................................................
    df['Rev_Dist'] = df['Run_Sum'] - df['Cum_Dist_Run']


#  Delete Short Run ........................................................................................
#     df = df[df['Run_No'] >= 10].reset_index(drop=True)

# Create a new column 'Pin_Point' with value "10 Meters" for the rows closest to 10 in each 'Run_No' group .........................................................
    # df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 10).idxmin()) == df.index, '10 Meters', '')
    # # Update 'Pin_Point' column with other values
    # df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 250).idxmin()) == df.index, '250 Meters', df['Pin_Point'])
    # df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 500).idxmin()) == df.index, '500 Meters', df['Pin_Point'])
    # df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 1000).idxmin()) == df.index, '1000 Meters', df['Pin_Point'])
    # Initialize the column
    df['Pin_Point'] = ''

    # Dictionary of target distances and their labels
    targets = {
        10: '10 Meters',
        250: '250 Meters',
        500: '500 Meters',
        1000: '1000 Meters'
    }

    # Loop through each target and label the closest point in each group
    for dist, label in targets.items():
        closest_indices = df.groupby(['Run_No', 'CMS_ID'])['Rev_Dist'].apply(lambda x: (x - dist).abs().idxmin())
        df.loc[closest_indices.values, 'Pin_Point'] = label

# Add BFT Column ......................................................................................................................
    df['Speed_shift'] = df['Speed'].shift(-1)
    unique_cms_ids = set()
    def add_bft(row):
        if row['Cum_Dist_LP'] < 10000 and 15 <= row['Speed'] <= 18 and row['Speed'] > row['Speed_shift']:
            if row['CMS_ID'] not in unique_cms_ids:
                unique_cms_ids.add(row['CMS_ID'])
                return 'BFT'
        return ''
    df['BFT'] = df.apply(add_bft, axis=1)

# Add BPT Column...............................................................................................................................
    unique_cms_ids = set()
    def add_bpt(row):
        if row['Cum_Dist_LP'] < 10000 and 40 <= row['Speed'] <= 90 and row['Speed'] > row['Speed_shift']:
            if row['CMS_ID'] not in unique_cms_ids:
                unique_cms_ids.add(row['CMS_ID'])
                return 'BPT'
        return ''
    df['BPT'] = df.apply(add_bpt, axis=1)

    # Reset shift column for comparison
    df['Speed_shift'] = df['Speed'].shift(1)

# BFT_END .............................................................................................................................
    def get_bft_end(df):
        df['BFT_END'] = ''
        for cms_id, group in df.groupby('CMS_ID'):
            group = group.reset_index()
            bft_idx = group[group['BFT'] == 'BFT'].index
            if not bft_idx.empty:
                bft_idx = bft_idx[0]
                for i in range(bft_idx + 1, len(group)):
                    if group.loc[i, 'Speed'] > group.loc[i, 'Speed_shift'] and 0 <= group.loc[i, 'Speed'] <= 9 and group.loc[i, 'Cum_Dist_LP'] < 10000:
                        end_idx = i - 1
                        if end_idx >= 0:
                            df.loc[group.loc[end_idx, 'index'], 'BFT_END'] = 'BFT_END'
                        break
        return df
# BPT_END ..............................................................................................................................
    def get_bpt_end(df):
        df['BPT_END'] = ''
        for cms_id, group in df.groupby('CMS_ID'):
            group = group.reset_index()
            bpt_idx = group[group['BPT'] == 'BPT'].index
            if not bpt_idx.empty:
                bpt_idx = bpt_idx[0]
                for i in range(bpt_idx + 1, len(group)):
                    if group.loc[i, 'Speed'] > group.loc[i, 'Speed_shift'] and 0 <= group.loc[i, 'Speed'] <= 45 and group.loc[i, 'Cum_Dist_LP'] < 10000:
                        end_idx = i - 1
                        if end_idx >= 0:
                            df.loc[group.loc[end_idx, 'index'], 'BPT_END'] = 'BPT_END'
                        break
        return df
    
    # Apply end marking functions
    df = get_bft_end(df)
    df = get_bpt_end(df)

    # Drop Speed_shift
    df.drop(columns=['Speed_shift'], inplace=True)

# Adding a new column 'BFT_BPT' ............................................................................................................................
    df['BFT_BPT'] = df.apply(lambda row: 
    (str(row['BFT']) if pd.notna(row['BFT']) and row['BFT'] != '' else '') +
    (' ' + str(row['BPT']) if pd.notna(row['BPT']) and row['BPT'] != '' else '') +
    (' ' + str(row['BFT_END']) if pd.notna(row['BFT_END']) and row['BFT_END'] != '' else '') +
    (' ' + str(row['BPT_END']) if pd.notna(row['BPT_END']) and row['BPT_END'] != '' else ''),
    axis=1
    ).str.strip()

    
 # Rearrange the Columns
    df = df[["Date", "Time", "Speed", "Distance", "CMS_ID", "Train_No", "Loco_No", "Cum_Dist_Run","Cum_Dist_LP","Run_No","Run_Sum","Rev_Dist","Pin_Point","BFT","BFT_END","BPT","BPT_END","BFT_BPT"]]
# Adding a new column 'Crew Name' , 'CLI Name', 'Desig'............................................................................................................................
    try:
        cms_file_path = 'CMS_Data.xlsx'
        cms_df = pd.read_excel(cms_file_path)

        # Your existing mapping logic using the CMS data
        df['Crew_Name'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[1]])
        df['Nom_CLI'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[4]])
        df['Desig'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[2]])

    except FileNotFoundError:
        # File not found handling: Create columns with null values
        df['Crew_Name'] = None
        df['Nom_CLI'] = None
        df['Desig'] = None

    print("Almost Done, Saving Sheet")
    window.text_box.clear()
    window.text_box.append("Saving File to Excel!!!")
    QApplication.processEvents()  # Force UI update
    # Save Data to Excel
    options = QFileDialog.Options() 
    save_path, _ = QFileDialog.getSaveFileName(None, "Save Processed Data", "", "Excel Files (*.xlsx)", options=options)

    # Check if a save location was selected
    if save_path:
        # Save the processed data to a single sheet "DataBase"
        df.to_excel(save_path, sheet_name="DataBase", index=False, engine="openpyxl")
        print("File saved successfully.")
        window.text_box.clear()
        window.text_box.append("Done Successfully!!!, Ready For Visual Analysis or Quick Report...")
        QApplication.processEvents()  # Force UI update
    else:
        print("Save canceled.")
        window.text_box.clear()
        window.text_box.append("Process Aborted...!!!!")
        QApplication.processEvents()  # Force UI update
    
    # print(df)
    return df



# =========== Laxvan =====================================================================================================
options = QFileDialog.Options()
def hide_group_box(group_box):
    if group_box:
        group_box.hide()
def laxvan_and_hide(cms_id_input,loco_no_input,train_no_input,process_button):
    print("CMS ID:", cms_id_input.text())
    print("Loco No:", loco_no_input.text())
    print("Train No:", train_no_input.text())
    
    process_button.hide()
    

    window.text_box.clear()
    window.text_box.append("Don't Click anywhere, Data is Processing .........")
    global options  # Access the global options variable
    if options is None:
        options = QFileDialog.Options()
    file_path, _ = QFileDialog.getOpenFileName(None, "Upload Medha File", "", "Text Files (*.txt);;Excel Files (*.xlsx *.xls)", options=options)
    if file_path:
        encodings_to_try = ['utf-8', 'latin-1', 'utf-16', 'windows-1252']
        print("Selected file:", file_path)

        df = None  # Initialize DataFrame

        try:
            if file_path.endswith('.txt'):
                for encoding in encodings_to_try:
                    try:
                        # Try reading the file while skipping metadata lines (first 5)
                        df = pd.read_csv(
                            file_path,
                            delimiter='\t',
                            encoding=encoding,
                            engine='python',
                            skiprows=5,
                            skip_blank_lines=True
                        )

                        # Basic validation
                        if df.shape[0] > 1 and df.shape[1] >= 4:
                            print(f"✅ Successfully read file with encoding {encoding}")
                            print(df.head())
                            break
                    except Exception as e:
                        print(f"❌ Failed with encoding {encoding}: {e}")
                else:
                    window.text_box.append("⚠️ Could not read the file with any encoding.")
                    return

            else:
                # For Excel files
                df = pd.read_excel(file_path)
                print("✅ Excel file loaded.")

            # ✅ Display or process the DataFrame
            window.text_box.clear()
            window.text_box.append("File loaded successfully.\n")
            window.text_box.append(str(df.head()))

        except Exception as e:
            window.text_box.clear()
            window.text_box.append(f"❌ An error occurred while processing the file: {str(e)}")

        # Calculate new Speed column based on Distance difference
    print(df.head())

    # Add computed Distance in meters
    df['Km'] = pd.to_numeric(df['Km'], errors='coerce')
    df['Distance'] = df['Km'].diff() * 1000
    df.loc[df.index[0], 'Distance'] = 0
    df['Distance'] = df['Distance'].round(2)

    # Step 3: Drop 'Km' and '---' columns if they exist
    columns_to_drop = ['Km', '---']
    df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True)


    try:
        first_col = df.columns[0]
        datetime_split = df[first_col].astype(str).str.strip().str.split(' ', n=1, expand=True)
        df['Date'] = datetime_split[0]
        df['Time'] = datetime_split[1]
        df.drop(columns=[first_col], inplace=True)

    except Exception as e:
        window.text_box.append(f"⚠️ Failed to split date and time: {e}")

    # Adding new columns with blank data
    df['CMS_ID'] = cms_id_input.text()
    df['Loco_No'] = loco_no_input.text()
    df['Train_No'] = train_no_input.text()
    
    df.rename(columns={'Km/hr': 'Speed'}, inplace=True)
    # Rearrange columns as "Date, Time, Speed, Distance, Train_No"............
    df = df[["Date", "Time", "Speed", "Distance", "CMS_ID", "Train_No", "Loco_No"]]

    print("Basic Columns Done")
    window.text_box.clear()
    window.text_box.append("Generating Excel- 25 %, Plase Wait..........")
    QApplication.processEvents()  # Force UI update

# .....Basic Columns Done..............................................................................................................................
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

# Convert certain columns to numeric format .................................................................................................
    numeric_columns = ['Speed', 'Distance', 'Loco_No']
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
    # Drop rows where "Speed" or "Distance" is NaN (non-numeric values)
    df = df.dropna(subset=["Speed", "Distance"])

    try:
        # Convert to datetime and reformat to DD/MM/YYYY
        df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%y', errors='coerce')
        df['Date'] = df['Date'].dt.strftime('%d/%m/%Y')  # Output format as string

    except Exception as e:
        window.text_box.append(f"⚠️ Failed to convert date format: {e}")

# Function to clean and convert the text to time format ..........................................................................................
    def clean_and_convert_time(time_str):
        time_str = time_str.strip()
        match = re.match(r'^(\d+):(\d+):(\d+)$', time_str)
        if match:
            hour, minute, second = match.groups()
            return f"{hour.zfill(2)}:{minute.zfill(2)}:{second.zfill(2)}"
        else:
            return None
    df['Time'] = df['Time'].apply(clean_and_convert_time)

# Cumulate the value of 'Distance' based on the Start & Stop .................................................................................
    # Convert 'Distance' column to numeric, coercing errors to NaN
    df['Distance'] = pd.to_numeric(df['Distance'], errors='coerce')
    df = df.dropna(subset=['Distance'])
    df['Cum_Dist_Run'] = df.groupby(df['Speed'].eq(0).cumsum())['Distance'].cumsum()

# Create a new column 'Run_No' based on the conditions ...................................................................................
    df['Run_No'] = 1  
    cms_change = df['CMS_ID'] != df['CMS_ID'].shift()
    distance_reset = df['Cum_Dist_Run'] < df['Cum_Dist_Run'].shift()
    group = cms_change.cumsum()
    df['Run_No'] = distance_reset.groupby(group).cumsum() + 1

# Cumulate the value of 'Distance' based on the Start & Stop .................................................................................
    df['Cum_Dist_Run'] = df.groupby(df['Speed'].eq(0).cumsum())['Distance'].cumsum()

# Cumulate the value of 'Distance' based on the LP .................................................................................
    if {'Speed', 'CMS_ID', 'Distance'}.issubset(df.columns):
        df['Cum_Dist_LP'] = df.groupby('CMS_ID')['Distance'].cumsum()

# Insert a new column 'Run_Sum' with the sum of 'Distance' per 'Run_No' ....................................................................
    df['Run_Sum'] = df.groupby(['Run_No','CMS_ID'])['Distance'].transform('sum')

# Deduct values from 'Run_Sum' to 'Distance' column ........................................................................................
    df['Rev_Dist'] = df['Run_Sum'] - df['Cum_Dist_Run']

# Create a new column 'Pin_Point' with value "10 Meters" for the rows closest to 10 in each 'Run_No' group .........................................................
    df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 10).idxmin()) == df.index, '10 Meters', '')
    # Update 'Pin_Point' column with other values
    df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 250).idxmin()) == df.index, '250 Meters', df['Pin_Point'])
    df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 500).idxmin()) == df.index, '500 Meters', df['Pin_Point'])
    df['Pin_Point'] = np.where(df.groupby(['Run_No','CMS_ID'])['Rev_Dist'].transform(lambda x: abs(x - 1000).idxmin()) == df.index, '1000 Meters', df['Pin_Point'])

# Add BFT Column ......................................................................................................................
    df['Speed_shift'] = df['Speed'].shift(-1)
    unique_cms_ids = set()
    def add_bft(row):
        if row['Cum_Dist_LP'] < 10000 and 15 <= row['Speed'] <= 18 and row['Speed'] > row['Speed_shift']:
            if row['CMS_ID'] not in unique_cms_ids:
                unique_cms_ids.add(row['CMS_ID'])
                return 'BFT'
        return ''
    df['BFT'] = df.apply(add_bft, axis=1)

# Add BPT Column...............................................................................................................................
    unique_cms_ids = set()
    def add_bpt(row):
        if row['Cum_Dist_LP'] < 10000 and 40 <= row['Speed'] <= 90 and row['Speed'] > row['Speed_shift']:
            if row['CMS_ID'] not in unique_cms_ids:
                unique_cms_ids.add(row['CMS_ID'])
                return 'BPT'
        return ''
    df['BPT'] = df.apply(add_bpt, axis=1)

    # Reset shift column for comparison
    df['Speed_shift'] = df['Speed'].shift(1)

# BFT_END .............................................................................................................................
    def get_bft_end(df):
        df['BFT_END'] = ''
        for cms_id, group in df.groupby('CMS_ID'):
            group = group.reset_index()
            bft_idx = group[group['BFT'] == 'BFT'].index
            if not bft_idx.empty:
                bft_idx = bft_idx[0]
                for i in range(bft_idx + 1, len(group)):
                    if group.loc[i, 'Speed'] > group.loc[i, 'Speed_shift'] and 0 <= group.loc[i, 'Speed'] <= 9 and group.loc[i, 'Cum_Dist_LP'] < 10000:
                        end_idx = i - 1
                        if end_idx >= 0:
                            df.loc[group.loc[end_idx, 'index'], 'BFT_END'] = 'BFT_END'
                        break
        return df
# BPT_END ..............................................................................................................................
    def get_bpt_end(df):
        df['BPT_END'] = ''
        for cms_id, group in df.groupby('CMS_ID'):
            group = group.reset_index()
            bpt_idx = group[group['BPT'] == 'BPT'].index
            if not bpt_idx.empty:
                bpt_idx = bpt_idx[0]
                for i in range(bpt_idx + 1, len(group)):
                    if group.loc[i, 'Speed'] > group.loc[i, 'Speed_shift'] and 0 <= group.loc[i, 'Speed'] <= 45 and group.loc[i, 'Cum_Dist_LP'] < 10000:
                        end_idx = i - 1
                        if end_idx >= 0:
                            df.loc[group.loc[end_idx, 'index'], 'BPT_END'] = 'BPT_END'
                        break
        return df
    
    # Apply end marking functions
    df = get_bft_end(df)
    df = get_bpt_end(df)

    # Drop Speed_shift
    df.drop(columns=['Speed_shift'], inplace=True)

# Adding a new column 'BFT_BPT' ............................................................................................................................
    df['BFT_BPT'] = df.apply(lambda row: 
    (str(row['BFT']) if pd.notna(row['BFT']) and row['BFT'] != '' else '') +
    (' ' + str(row['BPT']) if pd.notna(row['BPT']) and row['BPT'] != '' else '') +
    (' ' + str(row['BFT_END']) if pd.notna(row['BFT_END']) and row['BFT_END'] != '' else '') +
    (' ' + str(row['BPT_END']) if pd.notna(row['BPT_END']) and row['BPT_END'] != '' else ''),
    axis=1
    ).str.strip()

# Rearrange the Columns
    df = df[["Date", "Time", "Speed", "Distance", "CMS_ID", "Train_No", "Loco_No", "Cum_Dist_Run","Cum_Dist_LP","Run_No","Run_Sum","Rev_Dist","Pin_Point","BFT","BFT_END","BPT","BPT_END","BFT_BPT"]]
# Adding a new column 'Crew Name' , 'CLI Name', 'Desig'............................................................................................................................
    try:
        cms_file_path = 'CMS_Data.xlsx'
        cms_df = pd.read_excel(cms_file_path)
        # Your existing mapping logic using the CMS data
        df['Crew_Name'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[1]])
        df['Nom_CLI'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[4]])
        df['Desig'] = df['CMS_ID'].map(cms_df.set_index(cms_df.columns[0])[cms_df.columns[2]])

    except FileNotFoundError:
        # File not found handling: Create columns with null values
        df['Crew_Name'] = None
        df['Nom_CLI'] = None
        df['Desig'] = None
    print("Almost Done, Saving Sheet")
    window.text_box.clear()
    window.text_box.append("Saving File to Excel!!!")
    QApplication.processEvents()  # Force UI update
    # Save Data to Excel
    options = QFileDialog.Options() 
    save_path, _ = QFileDialog.getSaveFileName(None, "Save Processed Data", "", "Excel Files (*.xlsx)", options=options)

    # Check if a save location was selected
    if save_path:
        # Save the processed data to a single sheet "DataBase"
        df.to_excel(save_path, sheet_name="DataBase", index=False, engine="openpyxl")
        print("File saved successfully.")
        window.text_box.clear()
        window.text_box.append("Done Successfully!!!, Ready For Visual Analysis or Quick Report...")
        QApplication.processEvents()  # Force UI update
    else:
        print("Save canceled.")
        window.text_box.clear()
        window.text_box.append("Process Aborted...!!!!")
        QApplication.processEvents()  # Force UI update
    
    # print(df)
    return df   

def laxvan():
    window.text_box.clear()
    window.text_box.append("Fill Up Details & Upload Text File........")
    QApplication.processEvents()  # Force UI update
    # Create input fields for CMS ID, Loco No, and Train No
    group_box = QGroupBox("Input Fields *", window)
    group_box.setGeometry(120, 60, 380, 360)
    group_box.setStyleSheet("background-color: #8ffbf6  ")

    # Create input fields for CMS ID, Loco No, and Train No inside the group box
    cms_id_label = QLabel("CMS ID:", group_box)
    cms_id_label.setGeometry(30, 30, 100, 40)
    cms_id_input = QLineEdit(group_box)
    cms_id_input.setPlaceholderText("Only Uppercase")
    cms_id_input.setGeometry(140, 30, 200, 40)
    
    loco_no_label = QLabel("Loco No:", group_box)
    loco_no_label.setGeometry(30, 80, 100, 40)
    loco_no_input = QLineEdit(group_box)
    loco_no_input.setGeometry(140, 80, 200, 40)
    
    train_no_label = QLabel("Train No:", group_box)
    train_no_label.setGeometry(30, 130, 100, 40)
    train_no_input = QLineEdit(group_box)
    train_no_input.setGeometry(140, 130, 200, 40)
    
    process_button = QPushButton("Upload text File", group_box)
    process_button.setGeometry(110, 270, 200, 40)
    process_button.setStyleSheet("background-color: #025551; color: white; font-size: 20px;")

    # Instruction label
    instruction_label = QLabel("Warning:-Upload the .txt file for Single LP only", group_box)
    instruction_label.setGeometry(30, 200, 300, 60)
    instruction_label.setStyleSheet("color: red; font-size: 20px;")
    instruction_label.setWordWrap(True)
    
    # process_button.clicked.connect(partial(laxvan_and_hide, cms_id_input, train_no_input,loco_no_input, process_button))
    process_button.clicked.connect(partial(laxvan_and_hide, cms_id_input, loco_no_input, train_no_input, process_button))

    process_button.clicked.connect(lambda:hide_group_box(group_box))
    # process_button.clicked.connect(lambda: laxvan_and_hide(cms_id_input, loco_no_input, train_no_input))
    group_box.show()           

# =====QUICK REPORT=========================================================================================================================
def Quick_Report(self):
    window.text_box.clear()
    window.text_box.append("Select CMS ID from Uploaded File...")

    file_path, _ = QFileDialog.getOpenFileName(
        None, "Upload Formatted Excel File", "", "Excel Files (*.xlsx *.xls)", options=QFileDialog.Options()
    )
    if not file_path:
        return
    df = pd.read_excel(file_path, engine="openpyxl")

    if "CMS_ID" not in df.columns:
        window.text_box.clear()
        window.text_box.append("Error: 'CMS_ID' column missing from the uploaded file.")
        return
    unique_cms_ids = sorted(df["CMS_ID"].dropna().astype(str).unique())
    # Create a dropdown instead of a text box
    group_box = QGroupBox("Select CMS_ID *", window)
    group_box.setGeometry(120, 60, 400, 300)
    group_box.setStyleSheet("background-color: #DAF7A6;")

    cms_id_label = QLabel("CMS ID:", group_box)
    cms_id_label.setGeometry(30, 70, 100, 30)

    cms_id_dropdown = QComboBox(group_box)
    cms_id_dropdown.setGeometry(110, 70, 150, 30)
    cms_id_dropdown.addItems(unique_cms_ids)

    instruction_label = QLabel(
        "Select CMS ID from the list. Ensure that it is present in the processed file before uploading.", group_box
    )
    instruction_label.setGeometry(30, 120, 360, 80)
    instruction_label.setWordWrap(True)

    process_button = QPushButton("Generate Quick Report", group_box)
    process_button.setGeometry(100, 230, 230, 40)
    process_button.setStyleSheet("background-color: #05A7B4; color: white; font-size: 20px;")

    process_button.clicked.connect(partial(process_and_hide, cms_id_dropdown, group_box, df))
    
    group_box.show()

def process_and_hide(cms_id_dropdown, group_box, df):
    group_box.hide()
    window.text_box.clear()
    window.text_box.append("Processing data for selected CMS ID, please wait...")
    QApplication.processEvents()
    cms_id = cms_id_dropdown.currentText().strip()
    if not cms_id:
        window.text_box.clear()
        window.text_box.append("Error: Please select a CMS_ID before uploading.")
        QApplication.processEvents()
        return
    # Ensure df is a DataFrame and create a copy to avoid modifying original
    if not isinstance(df, pd.DataFrame):
        window.text_box.clear()
        window.text_box.append(f"Error: Input is not a DataFrame. Type: {type(df)}")
        QApplication.processEvents()
        return
    # Create a deep copy to prevent unintended modifications
    filtered_df = df[df["CMS_ID"] == cms_id].copy()
    # Validate DataFrame
    if filtered_df.empty:
        window.text_box.clear()
        window.text_box.append(f"Error: No records found for CMS_ID: {cms_id}")
        return
    # Ensure columns are of correct type
    column_type_conversions = {
        'Date': 'datetime64[ns]',
        'Time': 'datetime64[ns]',
        'Speed': 'float64',
        'Distance': 'float64',
        'Cum_Dist_LP': 'float64'
    }
    for col, dtype in column_type_conversions.items():
        if col in filtered_df.columns:
            try:
                filtered_df[col] = filtered_df[col].astype(dtype)
            except Exception as e:
                print(f"Warning: Could not convert {col} to {dtype}. Error: {e}")
    # Additional data validation
    required_columns = ['CMS_ID', 'Train_No', 'Loco_No', 'Desig', 'Crew_Name', 
                        'Nom_CLI', 'Distance', 'Speed', 'Run_No', 'Date', 'Time']
    
    missing_columns = [col for col in required_columns if col not in filtered_df.columns]
    if missing_columns:
        window.text_box.clear()
        window.text_box.append(f"Error: Missing columns: {', '.join(missing_columns)}")
        QApplication.processEvents()
        return
    try:
        # Convert Date and Time columns
        filtered_df['Date'] = pd.to_datetime(filtered_df['Date'], format='%d/%m/%Y', errors='coerce')
        filtered_df['Time'] = pd.to_datetime(filtered_df['Time'], format='%H:%M:%S', errors='coerce').dt.time
        filtered_df['DateTime'] = pd.to_datetime(
            filtered_df['Date'].dt.strftime('%d/%m/%Y') + ' ' + 
            filtered_df['Time'].astype(str), 
            format='%d/%m/%Y %H:%M:%S', 
            errors='coerce'
        )
        filtered_df = filtered_df.dropna(subset=['DateTime'])
        filtered_df = filtered_df.sort_values(by='DateTime')

    except Exception as e:
        window.text_box.clear()
        window.text_box.append(f"Error in datetime processing: {e}")
        QApplication.processEvents()
        return

    # Extract necessary details with error handling
    try:
        train_no = filtered_df["Train_No"].min()
        cms_id = filtered_df["CMS_ID"].min()
        loco_no = filtered_df["Loco_No"].min()
        designation = filtered_df["Desig"].iloc[0] if not filtered_df.empty else "N/A"
        pilot_name = filtered_df["Crew_Name"].unique()[0]
        nominated_cli = filtered_df["Nom_CLI"].unique()[0]
        total_km = round(filtered_df["Distance"].sum()/1000, 3)
        top_speed = filtered_df["Speed"].max()
        total_halt = filtered_df["Run_No"].max()
        # Ensure 'DateTime' column is in datetime format
        if "DateTime" in filtered_df.columns:
            filtered_df["DateTime"] = pd.to_datetime(filtered_df["DateTime"], errors="coerce")
        else:
            window.text_box.clear()
            window.text_box.append("Error: 'DateTime' column missing in DataFrame.")
            QApplication.processEvents()
            return
        # Ensure DataFrame is sorted before calculating time differences
        filtered_df = filtered_df.sort_values(by="DateTime")

        # Create 'TimeDiff' column if it doesn't exist
        if "TimeDiff" not in filtered_df.columns:
            filtered_df["TimeDiff"] = filtered_df["DateTime"].diff().dt.total_seconds().fillna(0)
        running_time_seconds = filtered_df.loc[filtered_df["Speed"] > 0, "TimeDiff"].sum()
        halt_time_seconds = filtered_df.loc[filtered_df["Speed"] == 0, "TimeDiff"].sum()
        running_time_seconds = max(0, int(running_time_seconds))
        halt_time_seconds = max(0, int(halt_time_seconds))
        running_time_str = str(pd.to_timedelta(running_time_seconds, unit="s"))
        halt_time_str = str(pd.to_timedelta(halt_time_seconds, unit="s"))
        total_time_seconds = running_time_seconds + halt_time_seconds
        avg_speed = round((total_km / total_time_seconds) * 3600, 2) if total_time_seconds > 0 else 0

        # Debugging print
        print(f"Running Time: {running_time_str}, Halt Time: {halt_time_str}, Avg Speed: {avg_speed}")

       # Convert min/max datetime strings back to datetime before subtraction
        min_datetime = pd.to_datetime(filtered_df["DateTime"].min())
        max_datetime = pd.to_datetime(filtered_df["DateTime"].max())
        total_duration = max_datetime - min_datetime
        # Format as string if needed
        min_datetime_str = min_datetime.strftime("%d-%m-%Y %H:%M:%S")
        max_datetime_str = max_datetime.strftime("%d-%m-%Y %H:%M:%S")
    except Exception as e:
        window.text_box.clear()
        window.text_box.append(f"Error extracting details: {e}")
        QApplication.processEvents()
        return

    # Print DataFrame details for debugging
    print("Filtered DataFrame Details:")
    print(filtered_df.info())

    # Call save_to_pdf with validated data
    save_to_pdf(
        cms_id, train_no, loco_no, total_km, 
        avg_speed, total_duration, designation, 
        pilot_name, total_halt, nominated_cli, 
        min_datetime_str, max_datetime_str, 
        filtered_df, running_time_str, 
        halt_time_str, top_speed
    )
# ===========================================================================================================================================
def save_to_pdf(cms_id, train_no, loco_no, total_distance , avg_speed, total_duration, designation, pilot_name , total_halt, nominated_cli, min_datetime_str, max_datetime_str, filtered_df, running_time_str, halt_time_str,top_speed):
# ===========================================================================================================================================    
    class CustomPDF(FPDF):
        def footer(self):
            """Footer method to add text at the bottom of each page."""
            # Position footer 15mm from bottom
            self.set_y(-15)
            # Add page number
            self.set_font("Arial", "I", 8)
            self.cell(0, 5, f"Page {self.page_no()}", align="C")
            # Add copyright and contact information
            self.set_y(-8)
            self.set_font("Arial", "I", 6)
            self.cell(0, 5, "Western Railway, Mumbai Division @BDTS1022 | Generated by SPM Analysis Tool", align="R")
# Initialize PDF using the custom class
    pdf = CustomPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.set_text_color(0, 0, 255)
    pdf.image("Logo.png", x=10, y=10, w=25, h=30)
    pdf.set_font("Arial", "B", 16)
    # Move cursor to the right of the image and align text
    pdf.set_xy(40, 15)  # Move text position to align with image
    pdf.cell(0, 10, "Loco Pilot Driving Technique Analysis", ln=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(40, 25)  # Align second line with the first one
    pdf.cell(0, 10, "Western Railway", ln=True, align="C")
    pdf.line(60, 35, 200, 35)
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 10, f"Report For Crew CMS ID: {cms_id}", ln=True, align="C")
    pdf.ln(10)
    pdf.cell(200, 10, f"Record Available For:- {min_datetime_str} to {max_datetime_str}", ln=True, align="L")
    pdf.cell(200, 10, f"Loco Pilot Name:- {pilot_name}", ln=True, align="L")
    pdf.cell(200, 10, f"Designation:- {designation}", ln=True, align="L")
    pdf.cell(200, 10, f"CMS ID:- {cms_id}", ln=True, align="L")
    pdf.cell(200, 10, f"Nominated CLI:- {nominated_cli}", ln=True, align="L")
    pdf.cell(200, 10, f"Loco Number:- {loco_no}", ln=True, align="L")
    pdf.cell(200, 10, f"Train Number:- {train_no}", ln=True, align="L")
    pdf.cell(200, 10, f"Start Date & Time:- {min_datetime_str}", ln=True, align="L")
    pdf.cell(200, 10, f"Finished Date & Time:- {max_datetime_str}", ln=True, align="L")
    pdf.cell(200, 10, f"Total Distance:- {total_distance} KM", ln=True)
    pdf.cell(200, 10, f"Wheel Start to Stop Duration:- {total_duration}  Hrs.", ln=True)
    pdf.cell(200, 10, f"Running Time:- {running_time_str}  Hrs.", ln=True)
    pdf.cell(200, 10, f"Halt Time:- {halt_time_str}  Hrs.", ln=True)
    pdf.cell(200, 10, f"Top-Speed:- {top_speed} Kmph", ln=True, align="L")
    pdf.cell(2050, 10, f"Avarage Speed:- {avg_speed} Kmph", ln=True, align="L")
    pdf.cell(200, 10, f"Total Halt:- {total_halt} Times", ln=True, align="L") 
    pdf.cell(200, 10, f"Total Distance:- {total_distance} KM", ln=True, align="L")
    pdf.ln(20)
    pdf.cell(200, 10, f"Prepared By :-  .............................................", ln=True, align="L") 
    pdf.ln(10)
    # print(filtered_df.dtypes)
    print("Text Page Done")
    window.text_box.clear()
    window.text_box.append("Page-1, Basic Data Page Created.........")
    QApplication.processEvents()
# =====Speed Slab Table & Data========================================================================================================================
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(20, 10, f"Speed Slab Wise Distance and Time", ln=True, align="L")       
   # Ensure 'Time' is in datetime format (FULL datetime, NOT just time)
    if not pd.api.types.is_datetime64_any_dtype(filtered_df["Time"]):
        filtered_df["Time"] = pd.to_datetime(filtered_df["Time"], format="%H:%M:%S", errors="coerce")
    filtered_df = filtered_df.dropna(subset=["DateTime"])
    filtered_df = filtered_df.sort_values(by="DateTime")
    filtered_df["TimeDiff"] = filtered_df["DateTime"].diff().fillna(pd.Timedelta(0))
    # Define speed slabs
    speed_slabs = [
        (1, 10), (11, 20), (21, 30), (31, 40), (41, 50),
        (51, 60), (61, 70), (71, 80), (81, 90), (91, 100),
        (101, 110), (111, 120), (121, 130), (131, 150)
    ]
    # Table Headers
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(60, 10, "Speed Slab (km/h)", border=1, align="C")
    pdf.cell(60, 10, "Total Distance (Meter)", border=1, align="C")
    pdf.cell(60, 10, "Total Time (HH:MM:SS)", border=1, align="C")
    pdf.ln()
    # Function to convert timedelta to "HH:MM:SS"
    def strfdelta(timedelta):
        if pd.isna(timedelta):  # Handle NaN cases
            return " "
        total_seconds = timedelta.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return f"{hours:02}:{minutes:02}:{seconds:02}"
    # Populate Table Data
    pdf.set_font("Arial", size=12)
    for start, end in speed_slabs:
        subset_df = filtered_df[(filtered_df['Speed'] >= start) & (filtered_df['Speed'] < end)]  # Use separate DataFrame
        total_distance = subset_df["Distance"].sum() / 1000 if not subset_df.empty else 0
        total_time = subset_df["TimeDiff"].sum() if not subset_df.empty else pd.Timedelta(0)

        total_distance_str = f"{total_distance:.3f} km" if total_distance > 0 else "- -"
        total_time_str = strfdelta(total_time) if total_time.total_seconds() > 0 else "- -"

        pdf.cell(60, 10, f"{start}-{end}", border=1, align="C")
        pdf.cell(60, 10, total_distance_str, border=1, align="C")
        pdf.cell(60, 10, total_time_str, border=1, align="C")
        pdf.ln()
    print("BPT Table Done")
    window.text_box.clear()
    window.text_box.append("Page-2, Speed Slab Table Created.........")
    QApplication.processEvents()
# =Bar Graph For Run Numbers===========================================================================================================
    pdf.add_page(orientation='L')  # Switch only this page to Landscape
    max_speed = filtered_df.groupby(['Run_No', 'CMS_ID'])['Speed'].max().reset_index()
    max_speed = max_speed[max_speed['CMS_ID'] == cms_id]
    sum_max_speed = max_speed.groupby('Run_No')['Speed'].sum().reset_index()
    max_sum_max_speed_index = sum_max_speed['Speed'].idxmax()
    fig = px.bar(sum_max_speed, x="Run_No", y="Speed", text="Speed")  # Add text="Speed"
    colors = ['#05B7B7'] * len(sum_max_speed)  # Light blue for most bars
    if max_sum_max_speed_index >= 0:
        colors[max_sum_max_speed_index] = '#854c03'  # Brown for the highest bar
    fig.update_traces(marker_color=colors)
    # Update Layout
    fig.update_layout(
        title=f"Max. Speed of Each Halt (CMS_ID: {cms_id})",
        xaxis_title="Halt Count",
        yaxis_title="Speed"
    )
    # Ensure X-axis is Categorical
    fig.update_xaxes(type='category')
    chart_path = "max_speed_chart.png"
    # Save Chart as Image
    pio.write_image(fig, chart_path, format="png", width=1200, height=600, scale=2)
    pdf.set_font("Arial", style="B", size=14)
    Loco_No = filtered_df["Loco_No"].min()
    pdf.cell(280, 10, f"Top Speed Between Each Halt (CMS_ID: {cms_id} & Loco_No {Loco_No})", ln=True, align='C')
    pdf.image(chart_path, x=10, y=20, w=285)  # Adjust width for landscape

    print("Bar Chart Done")
    window.text_box.clear()
    window.text_box.append("Page-3, Bar Chart Created........")
    QApplication.processEvents()

# Create Time VS Speed Area Graph in PDF instance in portrait mode  ================================================================================================== 
    filtered_cms_df = filtered_df[filtered_df["CMS_ID"] == cms_id].copy()
    filtered_cms_df["Time"] = pd.to_datetime(filtered_cms_df["Time"]).dt.strftime('%H:%M:%S')
    filtered_cms_df = filtered_cms_df.sort_values(by="Time")
    # Create Line Chart
    fig = px.line(
        filtered_cms_df,
        x="Time",
        y="Speed",
        title=f"Speed Variation Over Time for CMS_ID: {cms_id} Top Speed - {top_speed} Km/h",
        labels={"Time": "Time", "Speed": "Speed (km/h)"},
        color_discrete_sequence=["#3366CC"]
    )
    # Find Top Speed
    top_speed = filtered_cms_df["Speed"].max()
    top_speed_time = filtered_cms_df.loc[filtered_cms_df["Speed"] == top_speed, "Time"].iloc[0]

    # Improve Chart Appearance
    fig.update_layout(
        plot_bgcolor="white",
        xaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=14),
            tickangle=90,  # Rotate time labels for better visibility
            type="category",  # Treat time as categorical to prevent skipping values
            showline=True,  # Show Y-axis line

        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=14),
            showline=True,  # Show Y-axis line
            type="linear",
            layer="below traces",
            zeroline=True,  # Ensure zero line is visible
            zerolinecolor="blue",  # Make sure it's green
            zerolinewidth=1,  # Set thickness of the zero line   # Adjust thickness  # Adjust thickness
        ),
        hoverlabel=dict(
            bgcolor="white",
            font_size=12
        ),
        margin=dict(b=150)  # Extra margin for long time labels
    )
    # Add Top Speed Annotation
    fig.add_annotation(
        x=top_speed_time,
        y=top_speed,
        text="Top Speed",
        showarrow=True,
        arrowhead=4,
        arrowcolor="red",
        font=dict(color="black", size=12)
    )
    # Save Plotly Figure as an Image (PNG)
    graph_path = "speed_chart.png"
    pio.write_image(fig, graph_path, format="png", width=1000, height=500, scale=2)
    pdf.add_page(orientation='L')  # Only this page is Landscape
    pdf.set_font("Arial", style="B", size=14)
    Loco_No = filtered_cms_df["Loco_No"].min()  # Get Loco_No from filtered DataFrame
    pdf.cell(280, 10, f"Report for CMS_ID: {cms_id} & Loco_No {Loco_No}", ln=True, align='C')
    pdf.ln(10)  # Space before image
    pdf.image(graph_path, x=10, y=20, w=285)  # Adjust width for landscape

    print("Speed-Time Chart Done")
    window.text_box.clear()
    window.text_box.append("Page-4, Speed-Time Grpah Created.........")
    QApplication.processEvents()

# Create Distance VS Speed Area Graph in PDF instance in portrait mode  ....................................................................
    filtered_cms_df = filtered_df[filtered_df["CMS_ID"] == cms_id].copy()

    # Sort DataFrame by Distance Traveled
    filtered_cms_df = filtered_cms_df.sort_values(by="Cum_Dist_LP")

    # Define tick interval for X-Axis
    cum_dist_min = filtered_cms_df["Cum_Dist_LP"].min()
    cum_dist_max = filtered_cms_df["Cum_Dist_LP"].max()

    if pd.notna(cum_dist_min) and pd.notna(cum_dist_max):
        tick_step = max(1000, int((cum_dist_max - cum_dist_min) // 10) or 1)
        tick_values = list(range(int(cum_dist_min), int(cum_dist_max) + tick_step, tick_step))
    else:
        tick_values = []

    # Create Line Chart
    fig = px.line(
        filtered_cms_df,
        x="Cum_Dist_LP",
        y="Speed",
        title=f"Speed Variation Over Distance for CMS_ID: {cms_id}",
        labels={"Cum_Dist_LP": "Distance (km)", "Speed": "Speed (km/h)"},
        color_discrete_sequence=["#3366CC"]
    )

    # Find Top Speed
    top_speed = filtered_cms_df["Speed"].max()
    if not pd.isna(top_speed):
        top_speed_dist = filtered_cms_df.loc[filtered_cms_df["Speed"] == top_speed, "Cum_Dist_LP"].median()
    else:
        top_speed_dist = 0

    # Improve Chart Appearance
    fig.update_layout(
        plot_bgcolor="white",
        xaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=14),
            tickangle=90,
            tickmode="array",
            tickvals=tick_values,
            type="linear",
            showline=True,  # Show Y-axis line
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=14),
            showline=True,  # Show Y-axis line
            type="linear",
            layer="below traces",
            zeroline=True,  # Ensure zero line is visible
            zerolinecolor="blue",  # Make sure it's green
            zerolinewidth=1,  # Set thickness of the zero line   # Adjust thickness  # Adjust thickness
            
        ),
        hoverlabel=dict(
            bgcolor="white",
            font_size=12
        ),
        margin=dict(b=150)
    )

    # Add Top Speed Annotation
    fig.add_annotation(
        x=top_speed_dist,
        y=top_speed,
        text="Top Speed",
        showarrow=True,
        arrowhead=4,
        arrowcolor="red",
        font=dict(color="red", size=12)
    )

    # Save Plotly Figure as an Image (PNG)
    graph_path = "dist_chart.png"
    if not filtered_cms_df.empty:
        pio.write_image(fig, graph_path, format="png", width=1000, height=500, scale=2)
    else:
        print("Warning: No data available for graph generation.")

    # === Add Graph on Page 4 in Landscape Mode ===
    pdf.add_page(orientation='L')

    # Add title for Graph Page
    pdf.set_font("Arial", style="B", size=14)
    Loco_No = filtered_cms_df["Loco_No"].min()
    if pd.isna(Loco_No):
        Loco_No = "Unknown"

    pdf.cell(280, 10, f"Report for CMS_ID: {cms_id} & Loco_No {Loco_No}", ln=True, align='C')

    # Insert image in PDF
    pdf.ln(10)
    pdf.image(graph_path, x=10, y=20, w=285)
    print("Dist-Speed Chart Done")
    window.text_box.clear()
    window.text_box.append("Page-5, Distance-Speed Graph Created.........")
    QApplication.processEvents()


# BFT & BPT Table..............................................................................................
    filtered_cms_df["Time"] = pd.to_datetime(filtered_cms_df["Time"]).dt.strftime('%H:%M:%S')
    if "BPT" in filtered_df.columns:
        bpt_filtered = filtered_df[filtered_df["BPT"] == "BPT"]
    else:
        bpt_filtered = pd.DataFrame()
    pdf.add_page(orientation='L')  # Landscape Page
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(280, 10, f"Report for CMS_ID: {cms_id} & Loco_No {Loco_No}", ln=True, align='C')
    pdf.ln(30)
    pdf.set_font("Arial", style="", size=12)
    pdf.cell(280, 10, f"Brake Feel & Brake Power Test Conducted (Within 10 Km of First Section)", ln=True, align="L")
    pdf.ln(10)
    # Data for First Row - BFT Done
    bft_filtered = filtered_df[filtered_df["BFT"] == "BFT"]
    bft_time = (bft_filtered["Time"].iloc[0].strftime('%H:%M:%S') if not bft_filtered.empty else "N/A")
    bft_dist = bft_filtered["Cum_Dist_LP"].iloc[0] if not bft_filtered.empty else "N/A"
    bft_speed = bft_filtered["Speed"].iloc[0] if not bft_filtered.empty else "N/A"

    bft_end_filtered = filtered_df[filtered_df["BFT_END"] == "BFT_END"]
    bft_end_time = (bft_end_filtered["Time"].iloc[0].strftime('%H:%M:%S') if not bft_end_filtered.empty else "N/A")
    bft_end_dist = bft_end_filtered["Cum_Dist_LP"].iloc[0] if not bft_end_filtered.empty else "N/A"
    bft_end_speed = bft_end_filtered["Speed"].iloc[0] if not bft_end_filtered.empty else "N/A"

    try:
        bft_total_dist = float(bft_end_dist) - float(bft_dist)
    except (TypeError, ValueError):
        bft_total_dist = "Improper BFT"  # Handle cases where conversion fails


    # Data for Filter values where 'BPT' column contains 'BPT'
    bpt_filtered = filtered_df[filtered_df["BPT"] == "BPT"]
    bpt_time = (bpt_filtered["Time"].iloc[0].strftime('%H:%M:%S') if not bpt_filtered.empty else "N/A")
    bpt_dist = bpt_filtered["Cum_Dist_LP"].iloc[0] if not bpt_filtered.empty else "N/A"
    bpt_speed = bpt_filtered["Speed"].iloc[0] if not bpt_filtered.empty else "N/A"

    bpt_end_filtered = filtered_df[filtered_df["BPT_END"] == "BPT_END"]
    bpt_end_time = (bpt_end_filtered["Time"].iloc[0].strftime('%H:%M:%S') if not bpt_end_filtered.empty else "N/A")
    bpt_end_dist = bpt_end_filtered["Cum_Dist_LP"].iloc[0] if not bpt_end_filtered.empty else "N/A"
    bpt_end_speed = bpt_end_filtered["Speed"].iloc[0] if not bpt_end_filtered.empty else "N/A"

    try:
        bpt_total_dist = float(bpt_end_dist) - float(bpt_dist)
    except (TypeError, ValueError):
        bpt_total_dist = "Improper BPT"  # Handle cases where conversion fails
    # Create Table Header--------------------------------------------------------------------------------------
    pdf.set_font("Arial", style="B", size=10)
    pdf.cell(35, 10, "Test Done", border=1, align="C")
    pdf.cell(35, 10, "Start Time", border=1, align="C")
    pdf.cell(35, 10, "End Time", border=1, align="C")
    pdf.cell(35, 10, "Start Distance", border=1, align="C")
    pdf.cell(35, 10, "End Distance", border=1, align="C")
    pdf.cell(35, 10, "Total Distance", border=1, align="C")
    pdf.cell(35, 10, "Start Speed", border=1, align="C")
    pdf.cell(35, 10, "End Speed", border=1, align="C")
    pdf.ln()
    # Reset font for table content
    pdf.set_font("Arial", size=10)
   # First Row - BFT Done
    pdf.cell(35, 10, "BFT", border=1, align="C")
    pdf.cell(35, 10, str(bft_time), border=1, align="C")
    pdf.cell(35, 10, str(bft_end_time), border=1, align="C")
    pdf.cell(35, 10, str(bft_dist), border=1, align="C")
    pdf.cell(35, 10, str(bft_end_dist), border=1, align="C")
    pdf.cell(35, 10, str(bft_total_dist), border=1, align="C")
    pdf.cell(35, 10, str(bft_speed), border=1, align="C")
    pdf.cell(35, 10, str(bft_end_speed), border=1, align="C")
    pdf.ln()
    # Second Row - BPT Done
    pdf.cell(35, 10, "BPT", border=1, align="C")
    pdf.cell(35, 10, str(bpt_time), border=1, align="C")
    pdf.cell(35, 10, str(bpt_end_time), border=1, align="C")
    pdf.cell(35, 10, str(bpt_dist), border=1, align="C")
    pdf.cell(35, 10, str(bpt_end_dist), border=1, align="C")
    pdf.cell(35, 10, str(bpt_total_dist), border=1, align="C")
    pdf.cell(35, 10, str(bpt_speed), border=1, align="C")
    pdf.cell(35, 10, str(bpt_end_speed), border=1, align="C")
    pdf.ln(20)
    pdf.cell(280, 10, f"Note:- All Distances are Shown in Meters", ln=True, align="L")

# BPT BFT Done line Chart ..........................................................................................
    data_base = filtered_df[filtered_df['Cum_Dist_LP'] < 10000]
    bft = data_base[data_base['BFT_BPT'] == "BFT"]
    bpt = data_base[data_base['BFT_BPT'] == "BPT"]
    # Define tick interval for X-Axis
    tick_step = max(1000, (data_base["Cum_Dist_LP"].max() - data_base["Cum_Dist_LP"].min()) // 10)
    tick_values = list(range(int(data_base["Cum_Dist_LP"].min()), int(data_base["Cum_Dist_LP"].max()) + tick_step, tick_step))
    # Create Line Chart
    fig = px.line(
        data_base,
        x="Cum_Dist_LP",
        y="Speed",
        title="Brake Feel & Brake Power Test (In First Section of 10 KM)",
        labels={"Cum_Dist_LP": "Distance (mtr)", "Speed": "Speed (km/h)"},
        color_discrete_sequence=["#05B7B7"]
    )
    # Add text annotations at specific points
    Dist_points1 = bft.index.tolist() + bpt.index.tolist()
    text_values = ['BFT Done'] * len(bft) + ['BPT Done'] * len(bpt)
    for i, tp in enumerate(Dist_points1):
        if tp in data_base.index:
            selected_data = data_base.loc[tp]
            fig.add_annotation(
                x=selected_data['Cum_Dist_LP'],
                y=selected_data['Speed'],
                text=text_values[i],
                showarrow=True,
                arrowhead=4,
                arrowcolor="red",
                font=dict(color="black", size=12)
            )
    # Define X-axis ticks at multiples of 100 km
    tick_step = 100  # Fixed interval of 100 km
    tick_values = list(range(
        int(filtered_cms_df["Cum_Dist_LP"].min()), 
        int(filtered_cms_df["Cum_Dist_LP"].max()) + tick_step, 
        tick_step
    ))
    # Define Y-axis ticks at multiples of 100 km
    tick_step_y = 5  # Fixed interval of 100 km
    tick_values_y = list(range(
        int(filtered_cms_df["Speed"].min()), 
        int(filtered_cms_df["Speed"].max()) + tick_step, 
        tick_step_y
    ))
    # Improve Chart Appearance
    fig.update_layout(
        plot_bgcolor="white",
        xaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=9),
            tickangle=90,
            tickmode="array",
            tickvals=tick_values,
            type="linear",
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=9), 
            tickvals=tick_values_y,
            type="linear",
            layer="below traces",
            zeroline=True,  # Ensure zero line is visible
            zerolinecolor="blue",  # Make sure it's green
            zerolinewidth=1,  # Set thickness of the zero line   # Adjust thickness  # Adjust thickness
        ),
        hoverlabel=dict(
            bgcolor="white",
            font_size=12
        ),
        margin=dict(b=150)
    )
    # Save the figure as an image
    graph_path = "Bpt_graph.png"
    pio.write_image(fig, graph_path, format="png", width=1000, height=500, scale=2)
    pdf.add_page(orientation='L')
    pdf.set_font("Arial", style="B", size=14)
    Loco_No = data_base["Loco_No"].min()  # Get Loco_No from filtered DataFrame
    pdf.cell(280, 10, f"Report for CMS_ID: {cms_id} & Loco_No {Loco_No}", ln=True, align='C')
    # Insert image in PDF
    pdf.ln(10)  # Space before image
    pdf.image(graph_path, x=10, y=20, w=285)  # Adjust width for landscape

    print("BPT-BFT line Chart Done")
    window.text_box.clear()
    window.text_box.append("Page-6,BPT-BFT chart Created..........")
    QApplication.processEvents()
# Cumulative Graph ..............................................................................
    filtered_cms_df = filtered_df.sort_values(by="Time")
    start_time = filtered_cms_df["Time"].min()
    end_time = filtered_cms_df["Time"].max()
    time_series = pd.date_range(start=start_time, end=end_time, freq='60s')
    time_df = pd.DataFrame({"Time": time_series})
    time_df = time_df.sort_values(by="Time")
    merged_df = pd.merge_asof(time_df, filtered_cms_df[["Time", "Cum_Dist_LP"]], on="Time")
    merged_df["Cum_Dist_LP"] = merged_df["Cum_Dist_LP"].ffill()
    merged_df["Cum_Dist_LP"] = merged_df["Cum_Dist_LP"] / 1000
    merged_df["Time"] = merged_df["Time"].dt.strftime('%H:%M:%S')
    # Plot Graph
    fig = px.line(
        merged_df,
        x="Time",
        y="Cum_Dist_LP",
        title=f"Cumulative Distance Over Time for CMS_ID: {cms_id}",
        labels={"Time": "Time", "Cum_Dist_LP": "Cumulative Distance (KM))"},
        color_discrete_sequence=["#3366CC"]
    )
    # Update Graph Layout
    fig.update_layout(
        plot_bgcolor="white",
        xaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=12),
            tickangle=90,  # Rotate time labels
            type="category"
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="lightgray",
            title_font=dict(size=12),
            showline=True,  # Show Y-axis line
            type="linear",
            layer="below traces",
            zeroline=True,  # Ensure zero line is visible
            zerolinecolor="blue",  # Make sure it's green
            zerolinewidth=1,  # Set thickness of the zero line   # Adjust thickness  # Adjust thickness
        ),
        hoverlabel=dict(
            bgcolor="white",
            font_size=12
        ),
        margin=dict(b=150)  # Extra margin for labels
    )
    # Save Plot as Image
    graph_path = "cumulative_distance_chart.png"
    pio.write_image(fig, graph_path, format="png", width=1000, height=500, scale=2)
    pdf.add_page(orientation='L')  # Landscape Page
    # Add Title
    pdf.set_font("Arial", style="B", size=14)
    Loco_No = filtered_cms_df["Loco_No"].min()  # Get Loco_No
    pdf.cell(280, 10, f"Report for CMS_ID: {cms_id} & Loco_No {Loco_No}", ln=True, align='C')
    # Insert Image into PDF
    pdf.ln(10)  # Space before image
    pdf.image(graph_path, x=10, y=20, w=285)  # Adjust width for landscape

    print("Prograssive line Chart Done")
    window.text_box.clear()
    window.text_box.append("Page-7, Cumulative Speed Created..........")
    QApplication.processEvents()
# Run Number Wise Table......................................................................................................
    pdf.add_page(orientation='L')  # Switch to landscape for better table fit
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(280, 10, f"Run Details Table for CMS_ID: {cms_id}", ln=True, align='C')
    pdf.ln(5)
    # Create table headers
    pdf.set_font("Arial", style="B", size=10)
    col_widths = [20, 35, 35, 35, 35, 25, 25, 25, 25]  # Adjusted column widths for all columns including Top Speed
    headers = ["Run No", "Start Time", "End Time", "Distance (KM)", "Top Speed", "Speed@1000m", "Speed@500m", "Speed@350m", "Speed@10m"]
    
    # Print headers
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, align='C')
    pdf.ln()
    # Reset font for table content
    pdf.set_font("Arial", size=11)
    # Process data for each run
    run_numbers = sorted(filtered_df["Run_No"].unique())
    
    for run_no in run_numbers:
        try:
            run_data = filtered_df[filtered_df["Run_No"] == run_no].copy()
            start_time = run_data["Time"].min().strftime('%H:%M:%S') if not run_data.empty else "N/A"
            end_time = run_data["Time"].max().strftime('%H:%M:%S') if not run_data.empty else "N/A"
            distance = run_data["Distance"].sum()/1000
            top_speed = round(run_data["Speed"].max(), 2) if not run_data.empty else "N/A"
            def get_speed_at_point(point):
                try:
                    point_data = run_data[run_data["Pin_Point"] == point]
                    return round(point_data["Speed"].iloc[0], 2) if not point_data.empty else "N/A"
                except (IndexError, KeyError):
                    return "N/A"
            
            speed_1000m = get_speed_at_point("1000 Meters")
            speed_500m = get_speed_at_point("500 Meters")
            speed_250m = get_speed_at_point("250 Meters")
            speed_10m = get_speed_at_point("10 Meters")
            
            # Print row data
            row_data = [
                str(run_no),
                str(start_time),
                str(end_time),
                f"{distance:.2f}",
                f"{top_speed if top_speed != 'N/A' else 'N/A'}",
                f"{speed_1000m if speed_1000m != 'N/A' else 'N/A'}",
                f"{speed_500m if speed_500m != 'N/A' else 'N/A'}",
                f"{speed_250m if speed_250m != 'N/A' else 'N/A'}",
                f"{speed_10m if speed_10m != 'N/A' else 'N/A'}"
            ]
            
            for i, data in enumerate(row_data):
                pdf.cell(col_widths[i], 10, data, border=1, align='C')
            pdf.ln()
            
        except Exception as e:
            print(f"Error processing run {run_no}: {str(e)}")
            continue

    window.text_box.clear()
    window.text_box.append("Page-8, Run Details Table Created..........")
    QApplication.processEvents()

# Braking Pattern Chart ...............................................................................................
    braking_df = filtered_df[filtered_df["Rev_Dist"] < 1100].copy()
    valid_run_numbers = braking_df[braking_df["Run_Sum"] > 50]["Run_No"].unique()

    for run_no in valid_run_numbers:
        run_df = filtered_df[(filtered_df["Run_No"] == run_no) & (filtered_df["Rev_Dist"] < 1100)].copy()
        run_df = run_df.sort_values(by="Rev_Dist", ascending=False)
        top_speed = round(filtered_df[filtered_df["Run_No"] == run_no]["Speed"].max(),2)if not run_df.empty else "N/A"

        annotation_data = {
            1000: run_df[run_df["Pin_Point"] == "1000 Meters"],
            500: run_df[run_df["Pin_Point"] == "500 Meters"],
            250: run_df[run_df["Pin_Point"] == "250 Meters"],
            10: run_df[run_df["Pin_Point"] == "10 Meters"]
        }

        Dist_points = []
        Speed_points = []
        text_values = []

        for distance, data in annotation_data.items():
            if not data.empty:
                Dist_points.append(data["Rev_Dist"].min())
                Speed_points.append(data["Speed"].min())
                text_values.append(f"{distance} Meter<br>Speed: {data['Speed'].min()} kmph")

        # Plot Graph
        fig = px.line(
            run_df,
            x="Rev_Dist",
            y="Speed",
            title=f"Braking Pattern Before 1000 meter of Halt- {run_no} (Max. Speed of Run: {top_speed} kmph)",
            labels={"Rev_Dist": "Distance (in Meter)", "Speed": "Speed (kmph)"},
            color_discrete_sequence=["#FF5733"]
        )
        fig.update_layout(
            plot_bgcolor="white",
            xaxis=dict(
                showgrid=True,
                gridcolor="lightgray",
                title_font=dict(size=12),
                tickangle=90,
                type="linear",
                autorange="reversed"
            ),
            yaxis=dict(
                showgrid=True,
                gridcolor="lightgray",
                title_font=dict(size=12)
            ),
            hoverlabel=dict(
                bgcolor="white",
                font_size=12
            ),
            margin=dict(b=150),
        )

        # Add annotations
        for i in range(len(Dist_points)):
            fig.add_annotation(
                x=Dist_points[i],
                y=Speed_points[i],
                text=text_values[i],
                showarrow=True,
                arrowhead=5,
                arrowcolor='red'
            )
        graph_path = f"chart_run_{run_no}.png"
        pio.write_image(fig, graph_path, format="png", width=1200, height=600, scale=2)

        # Add a new page for each braking pattern
        pdf.add_page(orientation='L')
        pdf.set_font("Arial", style="B", size=14)
        pdf.cell(280, 10, f"Braking Pattern of : {cms_id}", ln=True, align='C')
        pdf.ln(10)
        pdf.image(graph_path, x=10, y=20, w=285)

        window.text_box.clear()
        window.text_box.append(f"Braking Pattern for Halt No- {run_no}")
        QApplication.processEvents()
    print("Braking Pattern Graph Done")


    file_path, _ = QFileDialog.getSaveFileName(None, "Save PDF File", f"Quick_Report_{loco_no}_{cms_id}.pdf", "PDF Files (*.pdf)")
    if not file_path:
        window.text_box.clear()
        window.text_box.append("PDF saving was cancelled.")
        QApplication.processEvents()  # Force UI update
        return
    pdf.output(file_path)
    window.text_box.clear()
    QApplication.processEvents()  # Force UI update
    window.text_box.append(f"PDF saved at {file_path}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = QMainWindow()
    window.setWindowTitle("LP Driving Skill Analysis-Western Railway")
    window.setWindowIcon(QIcon("main.png"))
    window.resize(1000, 820)

    # Create a QLabel to display the Loco image
    image_loco = QLabel(window)
    # Load the image using QPixmap and set it in the QLabel
    pixmap = QPixmap("image_1.png")  # Replace with the path to your image
    image_loco.setPixmap(pixmap)
    image_loco.setGeometry(35, 480, pixmap.width(), pixmap.height())  # Set the position and size

    # Create QLabel for the image
    image_label = QLabel(window)
    pixmap = QPixmap("spm.png")  # Use the image file name
    image_label.setPixmap(pixmap)
    image_label.setGeometry(45, 80, pixmap.width(), pixmap.height())
    # Version label
    version_label = QLabel('Version: 9.00 Lax_Quick', window)
    version_label.setGeometry(700, 10, 250, 60)
    version_label.setStyleSheet("font-weight: normal; color: grey; font-size: 18px;")

    button1 = QPushButton("Upload Medha .txt File", window)
    button1.setGeometry(560, 90, 350, 60)
    button1.setStyleSheet("background-color: #05A7B4; color: white; font-size: 25px;")
    button1.clicked.connect(medha)

    button6 = QPushButton("Upload Telpro .PDF File", window)
    button6.setGeometry(560, 180, 350, 60)
    button6.setStyleSheet("background-color: #05A7B4; color: white; font-size: 25px;")
    button6.clicked.connect(telpro)

    button3 = QPushButton("Process Laxvan .txt File", window)
    button3.setGeometry(560, 270, 350, 60)
    button3.setStyleSheet("background-color: #05A7B4; color: white; font-size: 25px;")
    button3.clicked.connect(laxvan)

    button4 = QPushButton("Visual Analysis", window)
    button4.setGeometry(560, 360, 350, 60)
    button4.setStyleSheet("background-color: #CD7F32; color: white; font-size: 25px;")
    button4.clicked.connect(launch_streamlit_app)

    button5 = QPushButton("Quick Report", window)
    button5.setGeometry(560, 445, 350, 60)
    button5.setStyleSheet("background-color: #080700; color: white; font-size: 25px;")
    button5.clicked.connect(Quick_Report)

    button7 = QPushButton("User Guide", window)
    button7.setGeometry(560, 535, 350, 60)
    button7.setStyleSheet("background-color: #026067; color: white; font-size: 25px;")
    button7.clicked.connect(FAQ)

    # Create a QTextEdit for displaying print commands
    window.text_box = QTextEdit(window)
    window.text_box.setGeometry(150, 700, 780, 100)
    window.text_box.setStyleSheet("background-color: #f0f0f0; color: blue; font-size: 22px; border: 2px solid #f0f0f0;")
    window.text_box.setAlignment(Qt.AlignCenter)

    
    window.show()
    sys.exit(app.exec_())

